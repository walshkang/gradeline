from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from grader.discovery import discover_submission_units, parse_folder_name


class DiscoveryTests(unittest.TestCase):
    def test_parse_folder_name(self) -> None:
        token, name = parse_folder_name("123-456 - Jane Doe - Feb 24, 2026 851 AM")
        self.assertEqual(token, "123-456")
        self.assertEqual(name, "Jane Doe")

    def test_discover_submission_units(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            student = root / "123-456 - Jane Doe - Feb 24, 2026 851 AM"
            student.mkdir()
            (student / "file1.pdf").write_bytes(b"%PDF-1.4")
            (student / "nested").mkdir()
            (student / "nested" / "file2.pdf").write_bytes(b"%PDF-1.4")
            (root / "no_pdfs").mkdir()

            units = discover_submission_units(root)
            self.assertEqual(len(units), 1)
            unit = units[0]
            self.assertEqual(unit.folder_token, "123-456")
            self.assertEqual(unit.student_name, "Jane Doe")
            self.assertEqual(len(unit.pdf_paths), 2)
            self.assertEqual(unit.pdf_paths[0].name, "file1.pdf")

    def test_discover_submission_units_skips_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            
            # A student folder with only index.html/htm/txt or hidden files
            student1 = root / "123 - Student One"
            student1.mkdir()
            (student1 / "index.html").write_text("html", encoding="utf-8")
            (student1 / "index.txt").write_text("text", encoding="utf-8")
            (student1 / ".hidden").write_text("hidden", encoding="utf-8")
            
            # A student folder with index.html AND a PDF
            student2 = root / "456 - Student Two"
            student2.mkdir()
            (student2 / "index.html").write_text("html", encoding="utf-8")
            (student2 / "submission.pdf").write_bytes(b"%PDF-1.4")
            
            # A file (not directory) at the root level of submissions
            (root / "index.html").write_text("html", encoding="utf-8")
            
            # A hidden folder at the root level of submissions
            (root / ".hidden_folder").mkdir()
            (root / ".hidden_folder" / "some.pdf").write_bytes(b"%PDF-1.4")
            
            units = discover_submission_units(root)
            
            # Only student2 should be discovered because student1 had no PDF/convertible files (index.html is ignored/skipped).
            self.assertEqual(len(units), 1)
            self.assertEqual(units[0].student_name, "Student Two")
            self.assertEqual(len(units[0].pdf_paths), 1)
            self.assertEqual(units[0].pdf_paths[0].name, "submission.pdf")

    def test_discover_submission_units_converts_xlsx_and_docx(self) -> None:
        import docx
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            student = root / "789 - Bella Mastendino"
            student.mkdir()

            # Create XLSX file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "HW3 Answers"
            ws["A1"] = "Question 1"
            ws["B1"] = "mu = 11.536"
            wb.save(student / "Isabella Mastendino.HW3.xlsx")

            # Create DOCX file
            doc = docx.Document()
            doc.add_paragraph("Homework 3 DOCX Submission")
            doc.save(student / "answers.docx")

            units = discover_submission_units(root)
            self.assertEqual(len(units), 1)
            unit = units[0]
            self.assertEqual(unit.student_name, "Bella Mastendino")

            # Check converted PDF paths
            pdf_names = {p.name for p in unit.pdf_paths}
            self.assertIn("Isabella Mastendino.HW3.pdf", pdf_names)
            self.assertIn("answers.pdf", pdf_names)

    def test_multi_file_submission_pdf_xlsx_heic(self) -> None:
        import openpyxl
        from PIL import Image

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            student = root / "999 - Multi Submission Student"
            student.mkdir()

            # 1. Existing PDF (template or main PDF)
            (student / "HW3.pdf").write_bytes(b"%PDF-1.4\n%EOF")

            # 2. XLSX file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "XLSX Solution"
            wb.save(student / "HW3.xlsx")

            # 3. Image file (PNG / JPEG)
            img = Image.new("RGB", (100, 100), color="blue")
            img.save(student / "solution_photo.png")

            units = discover_submission_units(root)
            self.assertEqual(len(units), 1)
            unit = units[0]

            pdf_names = {p.name for p in unit.pdf_paths}
            # All 3 files must be discovered (HW3.pdf, HW3_xlsx.pdf, 999 - Multi Submission Student_images.pdf)
            self.assertIn("HW3.pdf", pdf_names)
            self.assertIn("HW3_xlsx.pdf", pdf_names)
            self.assertIn("999 - Multi Submission Student_images.pdf", pdf_names)
            self.assertEqual(len(unit.pdf_paths), 3)

if __name__ == "__main__":
    unittest.main()

